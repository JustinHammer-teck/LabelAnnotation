"""
Additional services for Aviation Safety Annotation System
SchemaLoaderService and JSONExportService

These services should be integrated into services.py
"""
import logging
import json
from django.db import transaction

logger = logging.getLogger(__name__)


class SchemaLoaderService:
    """Load aviation annotation schema from 标签汇总.xlsx"""

    SHEET_CATEGORY_MAPPING = {
        '条目汇总': None,
        '基本信息': ['aircraft', 'event_labels', 'flight_phase'],
        '威胁类型&训练主题': 'threat',
        '威胁-管理&影响': ['threat_mgmt', 'threat_outcome'],
        '差错类型&训练主题': 'error',
        '差错-相关性&管理&影响': ['error_relevancy', 'error_mgmt', 'error_outcome'],
        'UAS&训练主题-UAS': 'uas',
        'UAS-相关性&管理': ['uas_relevancy', 'uas_mgmt'],
        '胜任力': 'competency',
        '训练评估': ['likelihood', 'severity', 'training_benefit'],
        'CRM训练主题': 'crm_topics',
    }

    def load_from_excel(self, excel_file_path):
        """Load all dropdown options from Excel file"""
        from .models import AviationDropdownOption
        from openpyxl import load_workbook
        import os

        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

        wb = load_workbook(excel_file_path, read_only=True, data_only=True)
        options_created = 0

        try:
            with transaction.atomic():
                deleted_count = AviationDropdownOption.objects.all().delete()[0]
                logger.info(f"Deleted {deleted_count} existing options")

                for sheet_name, categories in self.SHEET_CATEGORY_MAPPING.items():
                    if categories is None:
                        continue

                    if sheet_name not in wb.sheetnames:
                        logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                        continue

                    logger.info(f"Processing sheet: {sheet_name}")

                    if sheet_name == '基本信息':
                        count = self._load_basic_info(wb)
                    elif sheet_name in ['威胁类型&训练主题', '差错类型&训练主题', 'UAS&训练主题-UAS']:
                        count = self._load_hierarchical_with_training(wb, sheet_name, categories)
                    elif sheet_name in ['威胁-管理&影响', '差错-相关性&管理&影响', 'UAS-相关性&管理']:
                        count = self._load_management_outcome(wb, sheet_name, categories)
                    elif sheet_name == '胜任力':
                        count = self._load_competency(wb)
                    elif sheet_name == '训练评估':
                        count = self._load_training_evaluation(wb, categories)
                    elif sheet_name == 'CRM训练主题':
                        count = self._load_flat_list(wb, sheet_name, categories)
                    else:
                        count = 0

                    options_created += count
                    logger.info(f"  Created {count} options")

            logger.info(f"Successfully loaded {options_created} total options")
            return options_created

        finally:
            wb.close()

    def _load_basic_info(self, wb):
        """Load aircraft types, event labels, flight phases"""
        from .models import AviationDropdownOption

        ws = wb['基本信息']
        count = 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for idx, row in enumerate(rows):
            if not row or all(cell is None or cell == '' for cell in row):
                continue

            if row[0]:
                AviationDropdownOption.objects.update_or_create(
                    category='aircraft',
                    level=1,
                    code=f'AC{idx+1:03d}',
                    defaults={
                        'label': str(row[0]).strip(),
                        'display_order': idx,
                        'is_active': True
                    }
                )
                count += 1

            if len(row) > 1 and row[1]:
                AviationDropdownOption.objects.update_or_create(
                    category='event_labels',
                    level=1,
                    code=f'EL{idx+1:03d}',
                    defaults={
                        'label': str(row[1]).strip(),
                        'display_order': idx,
                        'is_active': True
                    }
                )
                count += 1

            if len(row) > 2 and row[2]:
                AviationDropdownOption.objects.update_or_create(
                    category='flight_phase',
                    level=1,
                    code=f'FP{idx+1:03d}',
                    defaults={
                        'label': str(row[2]).strip(),
                        'display_order': idx,
                        'is_active': True
                    }
                )
                count += 1

        return count

    def _load_hierarchical_with_training(self, wb, sheet_name, category):
        """Load 3-level hierarchical data with training topics"""
        from .models import AviationDropdownOption

        ws = wb[sheet_name]
        count = 0
        parent_l1 = None
        parent_l2 = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None or cell == '' for cell in row):
                continue

            l1_value = str(row[0]).strip() if row[0] else None
            l2_value = str(row[1]).strip() if len(row) > 1 and row[1] else None
            l3_value = str(row[2]).strip() if len(row) > 2 and row[2] else None
            training_topic = str(row[3]).strip() if len(row) > 3 and row[3] else None

            if l1_value:
                code, label = self._parse_code_label(l1_value)
                parent_l1, created = AviationDropdownOption.objects.update_or_create(
                    category=category,
                    level=1,
                    code=code,
                    defaults={
                        'label': label,
                        'parent': None,
                        'display_order': count,
                        'is_active': True
                    }
                )
                count += 1
                parent_l2 = None

            if l2_value:
                code, label = self._parse_code_label(l2_value)
                parent_l2, created = AviationDropdownOption.objects.update_or_create(
                    category=category,
                    level=2,
                    code=code,
                    defaults={
                        'label': label,
                        'parent': parent_l1,
                        'display_order': count,
                        'is_active': True
                    }
                )
                count += 1

            if l3_value:
                code, label = self._parse_code_label(l3_value)
                training_topics = [training_topic] if training_topic else []

                _, created = AviationDropdownOption.objects.update_or_create(
                    category=category,
                    level=3,
                    code=code,
                    defaults={
                        'label': label,
                        'parent': parent_l2 or parent_l1,
                        'training_topics': training_topics,
                        'display_order': count,
                        'is_active': True
                    }
                )
                count += 1

        return count

    def _load_management_outcome(self, wb, sheet_name, categories):
        """Load management and outcome options"""
        from .models import AviationDropdownOption

        ws = wb[sheet_name]
        count = 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for cat_idx, category in enumerate(categories):
            for idx, row in enumerate(rows):
                if not row or cat_idx >= len(row):
                    continue

                value = row[cat_idx]
                if value and str(value).strip():
                    AviationDropdownOption.objects.update_or_create(
                        category=category,
                        level=1,
                        code=f'{category.upper()[:3]}{idx+1:02d}',
                        defaults={
                            'label': str(value).strip(),
                            'display_order': idx,
                            'is_active': True
                        }
                    )
                    count += 1

        return count

    def _load_competency(self, wb):
        """Load competency indicators"""
        import re
        from .models import AviationDropdownOption

        ws = wb['胜任力']
        count = 0
        parent_l1 = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None or cell == '' for cell in row):
                continue

            l1_value = str(row[0]).strip() if row[0] else None
            l2_value = str(row[1]).strip() if len(row) > 1 and row[1] else None

            if l1_value:
                code, label = self._parse_competency_code(l1_value)
                parent_l1, created = AviationDropdownOption.objects.update_or_create(
                    category='competency',
                    level=1,
                    code=code,
                    defaults={
                        'label': label,
                        'parent': None,
                        'display_order': count,
                        'is_active': True
                    }
                )
                count += 1

            if l2_value:
                code, label = self._parse_code_label(l2_value)
                _, created = AviationDropdownOption.objects.update_or_create(
                    category='competency',
                    level=2,
                    code=code,
                    defaults={
                        'label': label,
                        'parent': parent_l1,
                        'display_order': count,
                        'is_active': True
                    }
                )
                count += 1

        return count

    def _parse_competency_code(self, value):
        """Parse competency L1 format: 'English (CODE)\\nChinese' -> (CODE, full_label)"""
        import re
        match = re.search(r'[(\（]([A-Z]{2,4})[)\）]', value)
        if match:
            code = match.group(1)
        else:
            code = value[:10]
        label = value.replace('\n', ' ')
        return code, label

    def _load_training_evaluation(self, wb, categories):
        """Load likelihood, severity, training benefit scales"""
        from .models import AviationDropdownOption

        ws = wb['训练评估']
        count = 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for cat_idx, category in enumerate(categories):
            for idx, row in enumerate(rows):
                if not row or cat_idx >= len(row):
                    continue

                value = row[cat_idx]
                if value and str(value).strip():
                    AviationDropdownOption.objects.update_or_create(
                        category=category,
                        level=1,
                        code=f'{category[:3].upper()}{idx+1}',
                        defaults={
                            'label': str(value).strip(),
                            'display_order': idx,
                            'is_active': True
                        }
                    )
                    count += 1

        return count

    def _load_flat_list(self, wb, sheet_name, category):
        """Load flat list of options"""
        from .models import AviationDropdownOption

        ws = wb[sheet_name]
        count = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or not row[0]:
                continue

            value = str(row[0]).strip()
            if value:
                AviationDropdownOption.objects.update_or_create(
                    category=category,
                    level=1,
                    code=f'{category[:3].upper()}{idx+1:02d}',
                    defaults={
                        'label': value,
                        'display_order': idx,
                        'is_active': True
                    }
                )
                count += 1

        return count

    def _parse_code_label(self, value):
        """Parse 'CODE label' format"""
        parts = value.split(maxsplit=2)
        if len(parts) >= 3:
            code = f'{parts[0]} {parts[1]}'
            label = parts[2]
        elif len(parts) == 2:
            code = parts[0]
            label = parts[1]
        else:
            code = value[:10]
            label = value

        return code, label

    def validate_excel_structure(self, excel_file_path):
        """Validate Excel file structure before loading"""
        from openpyxl import load_workbook
        import os

        if not os.path.exists(excel_file_path):
            return {'valid': False, 'errors': ['File not found']}

        errors = []

        try:
            wb = load_workbook(excel_file_path, read_only=True)
            expected_sheets = [s for s in self.SHEET_CATEGORY_MAPPING.keys() if self.SHEET_CATEGORY_MAPPING[s] is not None]
            actual_sheets = wb.sheetnames

            missing_sheets = set(expected_sheets) - set(actual_sheets)
            if missing_sheets:
                errors.append(f"Missing sheets: {missing_sheets}")

            wb.close()

            return {
                'valid': len(errors) == 0,
                'errors': errors
            }

        except Exception as e:
            return {
                'valid': False,
                'errors': [f'Failed to read Excel file: {str(e)}']
            }


class JSONExportService:
    """Fast JSON export service"""

    def export_to_json(self, project, format_type='flat'):
        """Export annotations to JSON"""
        from .models import AviationAnnotation

        annotations = AviationAnnotation.objects.filter(
            annotation__task__project=project
        ).select_related(
            'annotation__task__aviation_incident',
            'annotation__task',
            'annotation__completed_by'
        ).order_by('-created_at')

        if format_type == 'nested':
            data = [self._serialize_nested(ann) for ann in annotations]
        else:
            data = [self._serialize_flat(ann) for ann in annotations]

        return json.dumps(data, indent=2, ensure_ascii=False)

    def _serialize_flat(self, annotation):
        """Serialize annotation to flat format"""
        incident = getattr(annotation.annotation.task, 'aviation_incident', None)

        return {
            'event_number': incident.event_number if incident else '',
            'event_description': incident.event_description if incident else '',
            'date': str(incident.date) if incident else '',
            'time': str(incident.time) if incident and incident.time else '',
            'location': incident.location if incident else '',
            'airport': incident.airport if incident else '',
            'flight_phase': incident.flight_phase if incident else '',
            'aircraft_type': annotation.aircraft_type,
            'event_labels': annotation.event_labels,
            'threat_type_l1': annotation.threat_type_l1,
            'threat_type_l2': annotation.threat_type_l2,
            'threat_type_l3': annotation.threat_type_l3,
            'threat_management': annotation.threat_management,
            'threat_outcome': annotation.threat_outcome,
            'threat_description': annotation.threat_description,
            'threat_training_topics': annotation.threat_training_topics,
            'error_relevancy': annotation.error_relevancy,
            'error_type_l1': annotation.error_type_l1,
            'error_type_l2': annotation.error_type_l2,
            'error_type_l3': annotation.error_type_l3,
            'error_management': annotation.error_management,
            'error_outcome': annotation.error_outcome,
            'error_description': annotation.error_description,
            'error_training_topics': annotation.error_training_topics,
            'uas_relevancy': annotation.uas_relevancy,
            'uas_type_l1': annotation.uas_type_l1,
            'uas_type_l2': annotation.uas_type_l2,
            'uas_type_l3': annotation.uas_type_l3,
            'uas_management': annotation.uas_management,
            'uas_description': annotation.uas_description,
            'uas_training_topics': annotation.uas_training_topics,
            'competency_indicators': annotation.competency_indicators,
            'likelihood': annotation.likelihood,
            'severity': annotation.severity,
            'training_benefit': annotation.training_benefit,
            'competency_selections': annotation.competency_selections,
            'training_plan_ideas': annotation.training_plan_ideas,
            'goals_to_achieve': annotation.goals_to_achieve,
            'notes': annotation.notes,
            'annotated_by': annotation.annotation.completed_by.username if annotation.annotation.completed_by else '',
            'annotation_date': str(annotation.created_at),
        }

    def _serialize_nested(self, annotation):
        """Serialize annotation to nested format"""
        incident = getattr(annotation.annotation.task, 'aviation_incident', None)

        return {
            'incident': {
                'event_number': incident.event_number if incident else '',
                'event_description': incident.event_description if incident else '',
                'date': str(incident.date) if incident else '',
                'time': str(incident.time) if incident and incident.time else '',
                'location': incident.location if incident else '',
                'airport': incident.airport if incident else '',
                'flight_phase': incident.flight_phase if incident else '',
            },
            'annotation': {
                'basic_info': {
                    'aircraft_type': annotation.aircraft_type,
                    'event_labels': annotation.event_labels,
                },
                'threat': {
                    'type_l1': annotation.threat_type_l1,
                    'type_l2': annotation.threat_type_l2,
                    'type_l3': annotation.threat_type_l3,
                    'management': annotation.threat_management,
                    'outcome': annotation.threat_outcome,
                    'description': annotation.threat_description,
                    'training_topics': annotation.threat_training_topics,
                },
                'error': {
                    'relevancy': annotation.error_relevancy,
                    'type_l1': annotation.error_type_l1,
                    'type_l2': annotation.error_type_l2,
                    'type_l3': annotation.error_type_l3,
                    'management': annotation.error_management,
                    'outcome': annotation.error_outcome,
                    'description': annotation.error_description,
                    'training_topics': annotation.error_training_topics,
                },
                'uas': {
                    'relevancy': annotation.uas_relevancy,
                    'type_l1': annotation.uas_type_l1,
                    'type_l2': annotation.uas_type_l2,
                    'type_l3': annotation.uas_type_l3,
                    'management': annotation.uas_management,
                    'description': annotation.uas_description,
                    'training_topics': annotation.uas_training_topics,
                },
                'training': {
                    'competency_indicators': annotation.competency_indicators,
                    'likelihood': annotation.likelihood,
                    'severity': annotation.severity,
                    'training_benefit': annotation.training_benefit,
                    'competency_selections': annotation.competency_selections,
                    'training_plan_ideas': annotation.training_plan_ideas,
                    'goals_to_achieve': annotation.goals_to_achieve,
                },
            },
            'metadata': {
                'notes': annotation.notes,
                'annotated_by': annotation.annotation.completed_by.username if annotation.annotation.completed_by else '',
                'annotation_date': str(annotation.created_at),
            }
        }
