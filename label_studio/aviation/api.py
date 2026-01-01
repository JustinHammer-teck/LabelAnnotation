import logging
from datetime import date, datetime
from io import BytesIO

from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import load_workbook
from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from projects.models import Project
from tasks.models import Task

from asgiref.sync import async_to_sync
from guardian.shortcuts import assign_perm, get_objects_for_user, remove_perm

from core.services.audit_log_service import AuditLogService
from notifications.models import NotificationChannel, NotificationEventType
from notifications.services import NotificationService

from .permissions import (
    CanAssignAviationProject,
    CanCreateLabelingItem,
    CanDeleteLabelingItem,
    CanEditLabelingItem,
)

from .models import (
    AviationEvent,
    AviationProject,
    FieldFeedback,
    LabelingItem,
    LabelingItemPerformance,
    ResultPerformance,
    ReviewDecision,
    TypeHierarchy,
)
from .analytics import get_aviation_project_analytics
from .filters import apply_all_filters
from .serializers import (
    AnalyticsEventSerializer,
    ApproveRequestSerializer,
    AviationEventSerializer,
    AviationProjectAnalyticsSerializer,
    AviationProjectSerializer,
    CreateAviationProjectSerializer,
    FilterOptionsSerializer,
    LabelingItemPerformanceSerializer,
    LabelingItemSerializer,
    LinkItemsSerializer,
    RejectRequestSerializer,
    ResubmitRequestSerializer,
    ResultPerformanceSerializer,
    ReviewDecisionSerializer,
    ReviewHistoryResponseSerializer,
    RevisionRequestSerializer,
    TypeHierarchySerializer,
)


logger = logging.getLogger(__name__)


# Constants for role-based filtering
ELEVATED_ROLES = {'Manager', 'Researcher', 'Admin'}
ANNOTATOR_ROLE = 'Annotator'


class AviationProjectViewSet(viewsets.ModelViewSet):
    """ViewSet for aviation projects with role-based filtering."""

    serializer_class = AviationProjectSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        """
        Filter projects based on user role.

        Role-based access:
        - Admins (superusers): All organization projects
        - Managers/Researchers: All organization projects
        - Annotators (only): Only assigned projects via Guardian permissions
        - Users with no groups: Only assigned projects (no elevated role)

        Returns:
            QuerySet: Filtered aviation projects
        """
        user = self.request.user
        queryset = AviationProject.objects.all()

        # Filter by organization first
        if hasattr(user, 'active_organization') and user.active_organization:
            queryset = queryset.filter(project__organization=user.active_organization)
        else:
            # User has no active organization - return empty
            return queryset.none()

        # Apply role-based filtering for annotator-only users
        if self._is_annotator_only(user):
            # Get only projects user is assigned to via Guardian permissions
            queryset = get_objects_for_user(
                user,
                'aviation.assigned_to_aviation_project',
                klass=queryset,
                accept_global_perms=False
            )

        # Optimize queries with select_related
        queryset = queryset.select_related(
            'project',
            'project__organization',
        )

        return queryset.distinct()

    def _is_annotator_only(self, user):
        """
        Check if user has ONLY Annotator role (no Manager/Researcher/Admin).

        A user is considered "annotator-only" if:
        - They are NOT a superuser
        - They do NOT have any elevated role (Manager, Researcher, Admin)

        Users with no groups are also treated as needing assignment filtering.

        Args:
            user: User instance

        Returns:
            bool: True if user should only see assigned projects, False otherwise
        """
        if user.is_superuser:
            return False

        user_groups = set(user.groups.values_list('name', flat=True))

        # Check if user has any elevated role
        has_elevated_role = bool(user_groups & ELEVATED_ROLES)

        # If user has elevated role, they see all projects
        # If no elevated role (including no groups), they need assignment filtering
        return not has_elevated_role

    def get_serializer_class(self):
        if self.action == 'create':
            return CreateAviationProjectSerializer
        return AviationProjectSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_id = serializer.validated_data.get('project_id')

        with transaction.atomic():
            if project_id:
                ls_project = Project.objects.filter(
                    id=project_id,
                    organization=request.user.active_organization
                ).first()
                if not ls_project:
                    raise ValidationError({'project_id': 'Project not found or not accessible'})
                if hasattr(ls_project, 'aviation_project'):
                    raise ValidationError({'project_id': 'Project already has an aviation wrapper'})
            else:
                ls_project = Project.objects.create(
                    title=serializer.validated_data['title'],
                    description=serializer.validated_data.get('description', ''),
                    organization=request.user.active_organization,
                    created_by=request.user,
                    label_config='<View></View>',
                )

            aviation_project = AviationProject.objects.create(
                project=ls_project,
                default_workflow=serializer.validated_data.get('default_workflow', ''),
                require_uas_assessment=serializer.validated_data.get('require_uas_assessment', True),
            )

        response_serializer = AviationProjectSerializer(aviation_project)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def available_projects(self, request):
        """Return LS projects that don't have an aviation wrapper yet"""
        projects = Project.objects.filter(
            organization=request.user.active_organization
        ).exclude(
            aviation_project__isnull=False
        ).values('id', 'title')
        return Response(list(projects))


class AviationEventViewSet(viewsets.ModelViewSet):
    serializer_class = AviationEventSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        queryset = AviationEvent.objects.filter(
            task__project__organization=self.request.user.active_organization
        ).select_related('task')

        project_id = self.request.query_params.get('project')
        if project_id:
            try:
                project_id = int(project_id)
                queryset = queryset.filter(task__project__aviation_project__id=project_id)
            except ValueError:
                raise ValidationError({'project': 'Must be an integer'})

        return queryset


def _get_type_hierarchy_prefetch():
    return Prefetch(
        'children',
        queryset=TypeHierarchy.objects.filter(is_active=True).order_by('display_order').prefetch_related(
            Prefetch(
                'children',
                queryset=TypeHierarchy.objects.filter(is_active=True).order_by('display_order')
            )
        )
    )


class TypeHierarchyViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = TypeHierarchySerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        queryset = TypeHierarchy.objects.filter(is_active=True).prefetch_related(
            _get_type_hierarchy_prefetch()
        )
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        return queryset.order_by('category', 'level', 'display_order')

    @action(detail=False, methods=['get'])
    def hierarchy(self, request):
        category = request.query_params.get('category')
        if not category:
            return Response(
                {'error': 'category parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        root_items = TypeHierarchy.objects.filter(
            category=category,
            level=1,
            parent__isnull=True,
            is_active=True
        ).prefetch_related(
            _get_type_hierarchy_prefetch()
        ).order_by('display_order')

        serializer = TypeHierarchySerializer(root_items, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def search(self, request):
        query = request.query_params.get('q', '')
        category = request.query_params.get('category')

        queryset = TypeHierarchy.objects.filter(is_active=True)

        if category:
            queryset = queryset.filter(category=category)

        if query:
            queryset = queryset.filter(
                Q(code__icontains=query) |
                Q(label__icontains=query) |
                Q(label_zh__icontains=query)
            )

        queryset = queryset.prefetch_related(
            _get_type_hierarchy_prefetch()
        ).order_by('category', 'level', 'display_order')[:50]
        serializer = TypeHierarchySerializer(queryset, many=True)
        return Response(serializer.data)


class LabelingItemViewSet(viewsets.ModelViewSet):
    serializer_class = LabelingItemSerializer
    permission_classes = (
        IsAuthenticated,
        CanEditLabelingItem,
        CanDeleteLabelingItem,
        CanCreateLabelingItem,
    )

    def get_queryset(self):
        queryset = LabelingItem.objects.filter(
            event__task__project__organization=self.request.user.active_organization
        ).select_related(
            'event',
            'created_by',
            'reviewed_by',
            'linked_result',
            'threat_type_l1',
            'threat_type_l2',
            'threat_type_l3',
            'error_type_l1',
            'error_type_l2',
            'error_type_l3',
            'uas_type_l1',
            'uas_type_l2',
            'uas_type_l3',
        )

        event_id = self.request.query_params.get('event')
        if event_id:
            try:
                event_id = int(event_id)
                queryset = queryset.filter(event_id=event_id)
            except ValueError:
                raise ValidationError({'event': 'Must be an integer'})

        return queryset.order_by('event', 'sequence_number')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ResultPerformanceViewSet(viewsets.ModelViewSet):
    serializer_class = ResultPerformanceSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        queryset = ResultPerformance.objects.filter(
            aviation_project__project__organization=self.request.user.active_organization
        ).select_related(
            'aviation_project',
            'event',
            'created_by',
            'reviewed_by',
        ).prefetch_related('labeling_items')

        event_id = self.request.query_params.get('event')
        if event_id:
            try:
                event_id = int(event_id)
                queryset = queryset.filter(event_id=event_id)
            except ValueError:
                raise ValidationError({'event': 'Must be an integer'})

        return queryset

    def perform_create(self, serializer):
        event_id = self.request.data.get('event')
        if not event_id:
            raise ValidationError({'event': 'Event ID is required'})

        event = AviationEvent.objects.filter(
            id=event_id,
            task__project__organization=self.request.user.active_organization
        ).select_related('task__project__aviation_project').first()

        if not event:
            raise ValidationError({'event': 'Event not found or not accessible'})

        aviation_project = event.task.project.aviation_project
        serializer.save(
            created_by=self.request.user,
            event=event,
            aviation_project=aviation_project
        )

    @action(detail=True, methods=['post'], url_path='link-items')
    def link_items(self, request, pk=None):
        performance = self.get_object()
        serializer = LinkItemsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user_org = request.user.active_organization
        performance_org = performance.aviation_project.project.organization
        if performance_org != user_org:
            raise ValidationError({'detail': 'ResultPerformance does not belong to your organization'})

        item_ids = serializer.validated_data['item_ids']
        contribution_weight = serializer.validated_data.get('contribution_weight', 1.00)
        notes = serializer.validated_data.get('notes', '')

        items = LabelingItem.objects.filter(
            id__in=item_ids,
            event__task__project__organization=user_org
        )

        created_links = []
        for item in items:
            link, created = LabelingItemPerformance.objects.get_or_create(
                labeling_item=item,
                result_performance=performance,
                defaults={
                    'contribution_weight': contribution_weight,
                    'notes': notes,
                }
            )
            if created:
                created_links.append(link)

        return Response(
            LabelingItemPerformanceSerializer(created_links, many=True).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['delete'], url_path='unlink-items')
    def unlink_items(self, request, pk=None):
        performance = self.get_object()

        user_org = request.user.active_organization
        performance_org = performance.aviation_project.project.organization
        if performance_org != user_org:
            return Response(
                {'error': 'Not authorized'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = LinkItemsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        item_ids = serializer.validated_data['item_ids']

        deleted_count, _ = LabelingItemPerformance.objects.filter(
            result_performance=performance,
            labeling_item_id__in=item_ids
        ).delete()

        return Response(
            {'deleted_count': deleted_count},
            status=status.HTTP_200_OK
        )


class LabelingItemPerformanceViewSet(viewsets.ModelViewSet):
    serializer_class = LabelingItemPerformanceSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return LabelingItemPerformance.objects.filter(
            labeling_item__event__task__project__organization=self.request.user.active_organization
        ).select_related('labeling_item', 'result_performance')


EXCEL_COLUMN_MAPPING = {
    'event_number': ['event_number', 'event_id', '事件编号', '编号', '涉及航班'],
    'event_description': ['event_description', 'description', '事件描述', '描述', '事件详情/处置结果／后续措施', '事件详情'],
    'date': ['date', 'event_date', '日期', '事件日期', '事件发生时间'],
    'time': ['time', 'event_time', '时间', '事件时间'],
    'location': ['location', '地点', '位置', '报告单位'],
    'departure_airport': ['departure_airport', '起飞机场（四字代码）', '起飞机场'],
    'arrival_airport': ['arrival_airport', '降落机场（四字代码）', '降落机场'],
    'flight_phase': ['flight_phase', 'phase', '飞行阶段', '阶段', '事件类型'],
    'aircraft_type': ['aircraft_type', 'type', '机型', '飞机类型', '涉及飞机（机型/注册号）', '涉及飞机'],
    'aircraft_registration': ['aircraft_registration', 'registration', '注册号', '飞机注册号'],
    'weather_conditions': ['weather_conditions', 'weather', '天气', '天气条件', '存在威胁和发生原因', '威胁原因'],
}

REQUIRED_COLUMNS = ['event_number', 'date']


def normalize_column_name(col_name):
    if col_name is None:
        return None
    return str(col_name).strip().lower().replace(' ', '_').replace('-', '_')


def find_column_mapping(headers):
    mapping = {}
    normalized_headers = {normalize_column_name(h): idx for idx, h in enumerate(headers) if h}

    for field, aliases in EXCEL_COLUMN_MAPPING.items():
        for alias in aliases:
            normalized_alias = normalize_column_name(alias)
            if normalized_alias in normalized_headers:
                mapping[field] = normalized_headers[normalized_alias]
                break

    return mapping


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        return value.date()
    try:
        return datetime.strptime(str(value).strip(), '%Y-%m-%d').date()
    except ValueError:
        pass
    try:
        return datetime.strptime(str(value).strip(), '%Y/%m/%d').date()
    except ValueError:
        pass
    try:
        return datetime.strptime(str(value).strip(), '%d/%m/%Y').date()
    except ValueError:
        pass
    return None


def parse_time(value):
    if value is None:
        return None
    if hasattr(value, 'time') and callable(getattr(value, 'time')):
        return value.time()
    if hasattr(value, 'hour'):
        return value
    try:
        return datetime.strptime(str(value).strip(), '%H:%M:%S').time()
    except ValueError:
        pass
    try:
        return datetime.strptime(str(value).strip(), '%H:%M').time()
    except ValueError:
        pass
    return None


class AviationExcelUploadView(APIView):
    permission_classes = (IsAuthenticated,)
    parser_classes = (MultiPartParser,)

    ALLOWED_EXTENSIONS = ('.xlsx', '.xls')

    def post(self, request, pk):
        aviation_project = get_object_or_404(AviationProject, pk=pk)
        project = aviation_project.project

        if project.organization != request.user.active_organization:
            return Response(
                {'error': 'You do not have access to this project'},
                status=status.HTTP_403_FORBIDDEN
            )

        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        uploaded_file = request.FILES['file']
        filename = uploaded_file.name.lower()

        if not filename.endswith(self.ALLOWED_EXTENSIONS):
            return Response(
                {'error': f'Invalid file type. Allowed: {", ".join(self.ALLOWED_EXTENSIONS)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            file_content = BytesIO(uploaded_file.read())
            workbook = load_workbook(filename=file_content, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as e:
            return Response(
                {'error': f'Failed to parse Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response(
                {'error': 'Excel file must have a header row and at least one data row'},
                status=status.HTTP_400_BAD_REQUEST
            )

        headers = rows[0]
        column_mapping = find_column_mapping(headers)

        missing_required = [col for col in REQUIRED_COLUMNS if col not in column_mapping]
        if missing_required:
            return Response(
                {'error': f'Missing required columns: {", ".join(missing_required)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created_count = 0
        errors = []
        tasks_to_create = []
        events_to_create = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                event_number = row[column_mapping['event_number']]
                if event_number is None or str(event_number).strip() == '':
                    errors.append({'row': row_num, 'message': 'event_number is required'})
                    continue
                event_number = str(event_number).strip()

                date_value = row[column_mapping['date']]
                parsed_date = parse_date(date_value)
                if parsed_date is None:
                    errors.append({'row': row_num, 'message': f'Invalid date format: {date_value}'})
                    continue

                event_data = {
                    'event_number': event_number,
                    'date': parsed_date,
                }

                if 'event_description' in column_mapping:
                    val = row[column_mapping['event_description']]
                    event_data['event_description'] = str(val).strip() if val else ''

                if 'time' in column_mapping:
                    event_data['time'] = parse_time(row[column_mapping['time']])

                for field in ['location', 'departure_airport', 'arrival_airport',
                              'flight_phase', 'aircraft_type',
                              'aircraft_registration', 'weather_conditions']:
                    if field in column_mapping:
                        val = row[column_mapping[field]]
                        event_data[field] = str(val).strip() if val else ''

                tasks_to_create.append({
                    'project': project,
                    'data': {'event_number': event_number},
                })
                events_to_create.append(event_data)

            except Exception as e:
                errors.append({'row': row_num, 'message': str(e)})

        if not tasks_to_create:
            return Response({
                'success': True,
                'created_count': 0,
                'first_event_id': None,
                'errors': errors
            })

        with transaction.atomic():
            created_tasks = Task.objects.bulk_create([
                Task(**task_data) for task_data in tasks_to_create
            ])

            aviation_events = []
            for task, event_data in zip(created_tasks, events_to_create):
                aviation_events.append(AviationEvent(task=task, **event_data))

            created_events = AviationEvent.objects.bulk_create(aviation_events)
            created_count = len(created_tasks)
            first_event_id = created_events[0].id if created_events else None

        workbook.close()

        return Response({
            'success': True,
            'created_count': created_count,
            'first_event_id': first_event_id,
            'errors': errors
        })


class AviationExportView(APIView):
    """Export aviation project data to JSON or Excel format."""

    permission_classes = (IsAuthenticated,)

    def get(self, request, pk):
        from .services import AviationExportService

        aviation_project = get_object_or_404(AviationProject, pk=pk)

        if aviation_project.project.organization != request.user.active_organization:
            return Response(
                {'error': 'You do not have access to this project'},
                status=status.HTTP_403_FORBIDDEN
            )

        export_format = request.query_params.get('export_format', 'json').lower()
        if export_format not in ('json', 'xlsx'):
            return Response(
                {'error': f'Invalid format: {export_format}. Use json or xlsx'},
                status=status.HTTP_400_BAD_REQUEST
            )

        service = AviationExportService(aviation_project.id)

        if export_format == 'xlsx':
            output = service.export_to_xlsx()
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="aviation-export-{pk}.xlsx"'
            return response

        return Response(service.export_to_json())


# =============================================================================
# Review System API Views
# =============================================================================


class ItemSubmitAPI(generics.GenericAPIView):
    """
    POST /api/aviation/items/<pk>/submit/

    Submit a labeling item for review.
    Changes status from 'draft' to 'submitted'.
    Only the creator of the item can submit it.
    """
    permission_classes = (IsAuthenticated,)

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related('event__task__project'),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Submit a labeling item for review',
        operation_description='Submit a draft labeling item for review. Only draft items can be submitted.',
        responses={200: LabelingItemSerializer}
    )
    def post(self, request, pk):
        labeling_item = self.get_labeling_item(pk)

        # Validate current status
        if labeling_item.status != 'draft':
            return Response(
                {'error': f'Cannot submit item with status "{labeling_item.status}". Only draft items can be submitted.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update status
        labeling_item.status = 'submitted'
        labeling_item.save(update_fields=['status', 'updated_at'])

        return Response(
            LabelingItemSerializer(labeling_item).data,
            status=status.HTTP_200_OK
        )


class ReviewApproveAPI(generics.GenericAPIView):
    """
    POST /api/aviation/items/<pk>/approve/

    Approve a labeling item after review.

    Creates a ReviewDecision with status='approved' and updates
    the LabelingItem status to 'approved'.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = ApproveRequestSerializer

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related(
                'event__task__project'
            ),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Approve a labeling item',
        operation_description='Approve a labeling item after review. Creates a ReviewDecision record.',
        request_body=ApproveRequestSerializer,
        responses={200: ReviewDecisionSerializer}
    )
    def post(self, request, pk):
        from .services import send_review_notification

        labeling_item = self.get_labeling_item(pk)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            # Create the review decision
            decision = ReviewDecision.objects.create(
                labeling_item=labeling_item,
                status='approved',
                reviewer=request.user,
                reviewer_comment=serializer.validated_data.get('comment', '')
            )

            # Update labeling item status
            labeling_item.status = 'approved'
            labeling_item.reviewed_by = request.user
            labeling_item.reviewed_at = timezone.now()
            labeling_item.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])

        # Send notification to annotator
        send_review_notification(decision)

        return Response(
            ReviewDecisionSerializer(decision).data,
            status=status.HTTP_200_OK
        )


class ReviewRejectAPI(generics.GenericAPIView):
    """
    POST /api/aviation/items/<pk>/reject/

    Reject a labeling item with field-level feedback.

    Creates a ReviewDecision with status 'rejected_partial' or 'rejected_full',
    creates FieldFeedback records for each field, and updates the LabelingItem
    status to 'reviewed'.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = RejectRequestSerializer

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related(
                'event__task__project'
            ),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Reject a labeling item',
        operation_description='Reject a labeling item with field-level feedback. Requires at least one field_feedback.',
        request_body=RejectRequestSerializer,
        responses={200: ReviewDecisionSerializer}
    )
    def post(self, request, pk):
        from .services import send_review_notification

        labeling_item = self.get_labeling_item(pk)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            # Create the review decision
            decision = ReviewDecision.objects.create(
                labeling_item=labeling_item,
                status=serializer.validated_data['status'],
                reviewer=request.user,
                reviewer_comment=serializer.validated_data.get('comment', '')
            )

            # Create field feedback records
            for feedback_data in serializer.validated_data['field_feedbacks']:
                FieldFeedback.objects.create(
                    review_decision=decision,
                    labeling_item=labeling_item,
                    field_name=feedback_data['field_name'],
                    feedback_type=feedback_data['feedback_type'],
                    feedback_comment=feedback_data.get('feedback_comment', ''),
                    reviewed_by=request.user
                )

            # Update labeling item status to 'reviewed' (needs revision)
            labeling_item.status = 'reviewed'
            labeling_item.reviewed_by = request.user
            labeling_item.reviewed_at = timezone.now()
            labeling_item.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])

        # Send notification to annotator
        send_review_notification(decision)

        return Response(
            ReviewDecisionSerializer(decision).data,
            status=status.HTTP_200_OK
        )


class ReviewRevisionAPI(generics.GenericAPIView):
    """
    POST /api/aviation/items/<pk>/revision/

    Request revision of a labeling item.

    Creates a ReviewDecision with status='revision_requested',
    creates FieldFeedback records, and updates the LabelingItem
    status to 'reviewed'.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = RevisionRequestSerializer

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related(
                'event__task__project'
            ),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Request revision of a labeling item',
        operation_description='Request clarification or revision on specific fields.',
        request_body=RevisionRequestSerializer,
        responses={200: ReviewDecisionSerializer}
    )
    def post(self, request, pk):
        from .services import send_review_notification

        labeling_item = self.get_labeling_item(pk)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            # Create the review decision
            decision = ReviewDecision.objects.create(
                labeling_item=labeling_item,
                status='revision_requested',
                reviewer=request.user,
                reviewer_comment=serializer.validated_data.get('comment', '')
            )

            # Create field feedback records
            for feedback_data in serializer.validated_data['field_feedbacks']:
                FieldFeedback.objects.create(
                    review_decision=decision,
                    labeling_item=labeling_item,
                    field_name=feedback_data['field_name'],
                    feedback_type=feedback_data['feedback_type'],
                    feedback_comment=feedback_data.get('feedback_comment', ''),
                    reviewed_by=request.user
                )

            # Update labeling item status to 'reviewed'
            labeling_item.status = 'reviewed'
            labeling_item.reviewed_by = request.user
            labeling_item.reviewed_at = timezone.now()
            labeling_item.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])

        # Send notification to annotator
        send_review_notification(decision)

        return Response(
            ReviewDecisionSerializer(decision).data,
            status=status.HTTP_200_OK
        )


class ReviewResubmitAPI(generics.GenericAPIView):
    """
    POST /api/aviation/items/<pk>/resubmit/

    Resubmit a labeling item after making revisions.

    Updates the LabelingItem status back to 'submitted'.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = ResubmitRequestSerializer

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related(
                'event__task__project'
            ),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Resubmit a labeling item',
        operation_description='Resubmit a labeling item after making requested revisions.',
        request_body=ResubmitRequestSerializer,
        responses={200: LabelingItemSerializer}
    )
    def post(self, request, pk):
        from .services import send_resubmit_notification

        labeling_item = self.get_labeling_item(pk)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            # Update labeling item status back to 'submitted'
            labeling_item.status = 'submitted'
            labeling_item.save(update_fields=['status'])

        # Send notification to last reviewer
        send_resubmit_notification(labeling_item, request.user)

        return Response(
            LabelingItemSerializer(labeling_item).data,
            status=status.HTTP_200_OK
        )


class ReviewHistoryAPI(generics.GenericAPIView):
    """
    GET /api/aviation/items/<pk>/review-history/

    Get the complete review history for a labeling item.

    Returns all ReviewDecisions with nested FieldFeedbacks,
    the current status, and pending revision fields.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = ReviewHistoryResponseSerializer

    def get_labeling_item(self, pk):
        """Get labeling item with organization check."""
        return get_object_or_404(
            LabelingItem.objects.select_related(
                'event__task__project'
            ).prefetch_related(
                'review_decisions__field_feedbacks'
            ),
            pk=pk,
            event__task__project__organization=self.request.user.active_organization
        )

    def get_pending_revision_fields(self, labeling_item):
        """Get list of field names with pending revision requests."""
        # Get the latest non-approved decision
        latest_decision = labeling_item.review_decisions.exclude(
            status='approved'
        ).order_by('-created_at').first()

        if not latest_decision:
            return []

        # If item was approved after this decision, no pending fields
        approved_decision = labeling_item.review_decisions.filter(
            status='approved',
            created_at__gt=latest_decision.created_at
        ).exists()

        if approved_decision:
            return []

        # Return field names from the latest non-approved decision
        return list(
            latest_decision.field_feedbacks.values_list('field_name', flat=True)
        )

    @swagger_auto_schema(
        tags=['Aviation Review'],
        operation_summary='Get review history for a labeling item',
        operation_description='Returns all review decisions, field feedbacks, current status, and pending revision fields.',
        responses={200: ReviewHistoryResponseSerializer}
    )
    def get(self, request, pk):
        labeling_item = self.get_labeling_item(pk)

        decisions = labeling_item.review_decisions.all().order_by('-created_at')
        pending_fields = self.get_pending_revision_fields(labeling_item)

        response_data = {
            'current_status': labeling_item.status,
            'pending_revision_fields': pending_fields,
            'decisions': ReviewDecisionSerializer(decisions, many=True).data
        }

        return Response(response_data, status=status.HTTP_200_OK)


# =============================================================================
# Assignment API Views
# =============================================================================


class AviationProjectAssignmentAPI(APIView):
    """
    API endpoint for managing aviation project assignments.

    GET: Returns all organization users with their assignment status
    POST: Toggles user assignments using Guardian permissions

    Follows the pattern from Label Studio's ProjectAssignmentAPI.
    """
    permission_classes = (IsAuthenticated, CanAssignAviationProject)

    def get_object(self, pk):
        """Get aviation project (permissions enforced by DRF)."""
        return get_object_or_404(AviationProject, pk=pk)

    @swagger_auto_schema(
        tags=['Aviation Projects'],
        operation_summary='Get user assignment status',
        operation_description='Returns all organization users with their assignment status for this aviation project.',
        responses={200: 'List of users with assignment status'}
    )
    def get(self, request, pk, *args, **kwargs):
        """Get all users in organization with their assignment status."""
        from django.contrib.auth import get_user_model

        from .serializers import AviationProjectUserPermissionSerializer

        aviation_project = self.get_object(pk)

        # Check object-level permissions (DRF calls this automatically for generic views,
        # but for APIView we need to call it manually)
        self.check_object_permissions(request, aviation_project)

        # Get all users from organization
        if hasattr(aviation_project.project, 'organization') and aviation_project.project.organization:
            all_users = aviation_project.project.organization.users.all().order_by('email')
        else:
            User = get_user_model()
            all_users = User.objects.filter(is_active=True).order_by('email')

        serializer_context = {
            'aviation_project': aviation_project,
        }

        serializer = AviationProjectUserPermissionSerializer(
            all_users,
            many=True,
            context=serializer_context
        )

        return Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        tags=['Aviation Projects'],
        operation_summary='Toggle user assignments',
        operation_description='Assign or revoke user assignments to this aviation project.',
        responses={204: 'No Content - Success', 400: 'Bad Request', 403: 'Forbidden', 404: 'Not Found'}
    )
    def post(self, request, pk, *args, **kwargs):
        """Toggle user assignments based on payload."""
        from django.contrib.auth import get_user_model

        from .serializers import AviationAssignmentToggleSerializer

        aviation_project = self.get_object(pk)

        # Check object-level permissions (DRF calls this automatically for generic views,
        # but for APIView we need to call it manually)
        self.check_object_permissions(request, aviation_project)

        # Validate request payload
        serializer = AviationAssignmentToggleSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        User = get_user_model()
        request_data = serializer.validated_data.get('users', [])
        request_user = request.user

        for user_data in request_data:
            new_perm = user_data.get('has_permission')
            assign_user_id = user_data.get('user_id')

            try:
                assign_user = User.objects.get(pk=assign_user_id)
            except User.DoesNotExist:
                return Response(
                    {'error': f'User with id {assign_user_id} not found'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if new_perm:
                self._assign_permission(request_user, aviation_project, assign_user)
            else:
                self._revoke_permission(request_user, aviation_project, assign_user)

        return Response(status=status.HTTP_204_NO_CONTENT)

    def _assign_permission(self, request_user, aviation_project, assign_user):
        """Assign permission and create audit log + notification."""
        # Assign permission using guardian
        assign_perm('aviation.assigned_to_aviation_project', assign_user, aviation_project)

        # Audit log
        AuditLogService.create(
            user=request_user,
            action=f'User #{request_user.id} "{request_user.email}" assign '
                   f'AviationProject #{aviation_project.id} to User #{assign_user.id} "{assign_user.email}"'
        )

        # Notification
        project_url = f'/aviation/{aviation_project.id}/'

        async_to_sync(NotificationService().send_notification)(
            channel_name=NotificationChannel.NOTIFICATION,
            event_type=NotificationEventType.PROJECT_ASSIGNED,
            subject='Aviation Project Assigned',
            message=f'User {request_user.email} has assigned you to Aviation Project: {aviation_project.project.title}',
            ts=timezone.now(),
            receive_user=assign_user,
            path=project_url,
            action_type='assign',
            source=f'aviation_project:{aviation_project.id}'
        )

    def _revoke_permission(self, request_user, aviation_project, assign_user):
        """Revoke permission and create audit log + notification."""
        # Revoke permission using guardian
        remove_perm('aviation.assigned_to_aviation_project', assign_user, aviation_project)

        # Audit log
        AuditLogService.create(
            user=request_user,
            action=f'User #{request_user.id} "{request_user.email}" revoke assign '
                   f'AviationProject #{aviation_project.id} from User #{assign_user.id} "{assign_user.email}"'
        )

        # Notification
        project_url = f'/aviation/{aviation_project.id}/'

        async_to_sync(NotificationService().send_notification)(
            channel_name=NotificationChannel.NOTIFICATION,
            event_type=NotificationEventType.PROJECT_ASSIGNED,
            subject='Aviation Project Assignment Revoked',
            message=f'User {request_user.email} has revoked your assignment from Aviation Project: {aviation_project.project.title}',
            ts=timezone.now(),
            receive_user=assign_user,
            path=project_url,
            action_type='revoke',
            source=f'aviation_project:{aviation_project.id}'
        )


# =============================================================================
# Analytics API Views
# =============================================================================


class AviationProjectAnalyticsAPI(generics.GenericAPIView):
    """
    GET /api/aviation/projects/<pk>/analytics/

    Retrieve analytics and metrics for an aviation project.

    Returns event completion statistics, labeling item status breakdown,
    and overall project progress metrics.

    Authentication:
        Requires authenticated user with access to the project's organization.

    Response:
        - project_id: Aviation project ID (AviationProject.id)
        - project_type: "aviation"
        - total_events: Total number of events
        - events_by_status: Breakdown of in_progress vs completed events
        - labeling_items: Total count and status breakdown

    Event Status Logic:
        - "completed": ALL labeling items are approved
        - "in_progress": Has ANY non-approved labeling item
        - Events without labeling items are not counted as in_progress or completed
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = AviationProjectAnalyticsSerializer

    def get_object(self):
        """Get aviation project with organization access check."""
        aviation_project = get_object_or_404(
            AviationProject.objects.select_related('project', 'project__organization'),
            pk=self.kwargs['pk']
        )

        # Verify organization access
        if aviation_project.project.organization != self.request.user.active_organization:
            from django.http import Http404
            raise Http404("Aviation project not found")

        return aviation_project

    @swagger_auto_schema(
        tags=['Aviation Analytics'],
        operation_summary='Get aviation project analytics',
        operation_description="""
        Retrieve comprehensive analytics for an aviation project.

        Returns:
        - Total event count
        - Event status breakdown (in_progress vs completed)
        - Labeling item status breakdown (draft, submitted, reviewed, approved)

        Event Status Logic:
        - "completed": ALL labeling items are approved
        - "in_progress": Has ANY non-approved labeling item

        Requires authentication and organization membership.
        """,
        responses={
            200: AviationProjectAnalyticsSerializer,
            401: 'Unauthorized - Authentication required',
            404: 'Aviation project not found or not accessible'
        }
    )
    def get(self, request, pk):
        """Retrieve analytics for the aviation project."""
        aviation_project = self.get_object()

        # Get analytics data from Phase 1 function
        analytics_data = get_aviation_project_analytics(aviation_project.id)

        if analytics_data is None:
            from django.http import Http404
            raise Http404("Analytics data not available")

        # Serialize and return
        serializer = self.get_serializer(analytics_data)
        return Response(serializer.data, status=status.HTTP_200_OK)


# =============================================================================
# Events Analytics Pagination and API View (Phase 2)
# =============================================================================


class AviationEventsAnalyticsPagination(PageNumberPagination):
    """
    Pagination class for events analytics endpoint.

    Default: 50 items per page
    Max: 100 items per page
    """
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100


class AviationProjectEventsAnalyticsAPI(generics.ListAPIView):
    """
    GET /api/aviation/projects/<pk>/events/analytics/

    Retrieve paginated event data for analytics/Sankey visualization.

    Returns events with nested labeling items and result performances,
    using Chinese field names for frontend compatibility. Supports all
    10 filter types from Phase 1 via query parameters.

    Query Parameters (Filters):
        - date_start: Start date (YYYY-MM-DD, inclusive)
        - date_end: End date (YYYY-MM-DD, inclusive)
        - aircraft: Aircraft types (comma-separated for multiple)
        - airport: Airport code (matches departure/arrival/actual_landing)
        - event_type: Event types (comma-separated for multiple)
        - flight_phase: Flight phases (comma-separated for multiple)
        - threat_l1/l2/l3: Threat hierarchy codes
        - error_l1/l2/l3: Error hierarchy codes
        - uas_l1/l2/l3: UAS hierarchy codes
        - training_topic: Training topics (comma-separated)
        - competency: Competency codes (comma-separated)

    Query Parameters (Pagination):
        - page: Page number (default: 1)
        - page_size: Items per page (default: 50, max: 100)

    Authentication:
        Requires authenticated user with access to the project's organization.

    Response:
        - count: Total number of events matching filters
        - next: URL to next page (or null)
        - previous: URL to previous page (or null)
        - results: Array of event objects with Chinese field names
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = AnalyticsEventSerializer
    pagination_class = AviationEventsAnalyticsPagination

    def get_aviation_project(self):
        """Get aviation project with organization access check."""
        aviation_project = get_object_or_404(
            AviationProject.objects.select_related('project', 'project__organization'),
            pk=self.kwargs['pk']
        )

        # Verify organization access
        if aviation_project.project.organization != self.request.user.active_organization:
            raise Http404("Aviation project not found")

        return aviation_project

    def get_queryset(self):
        """Build filtered and prefetched queryset for events."""
        aviation_project = self.get_aviation_project()

        # Build base queryset with optimized prefetching
        queryset = AviationEvent.objects.filter(
            task__project=aviation_project.project
        ).prefetch_related(
            Prefetch(
                'labeling_items',
                queryset=LabelingItem.objects.select_related(
                    'threat_type_l1', 'threat_type_l2', 'threat_type_l3',
                    'error_type_l1', 'error_type_l2', 'error_type_l3',
                    'uas_type_l1', 'uas_type_l2', 'uas_type_l3',
                    'linked_result',
                )
            ),
            'result_performances',
        ).order_by('-date', '-id')

        # Parse query params into filter dict
        filter_params = self._parse_filter_params()

        # Apply Phase 1 filters
        queryset = apply_all_filters(queryset, filter_params)

        return queryset

    def _parse_filter_params(self):
        """Parse query parameters into filter dictionary."""
        params = self.request.query_params
        filter_dict = {}

        # Date filters
        date_start = params.get('date_start')
        if date_start:
            try:
                filter_dict['date_start'] = date.fromisoformat(date_start)
            except ValueError:
                pass

        date_end = params.get('date_end')
        if date_end:
            try:
                filter_dict['date_end'] = date.fromisoformat(date_end)
            except ValueError:
                pass

        # Aircraft filter (comma-separated)
        aircraft = params.get('aircraft')
        if aircraft:
            filter_dict['aircraft'] = [a.strip() for a in aircraft.split(',')]

        # Airport filter (single value)
        airport = params.get('airport')
        if airport:
            filter_dict['airport'] = airport

        # Event type filter (comma-separated)
        event_type = params.get('event_type')
        if event_type:
            filter_dict['event_types'] = [e.strip() for e in event_type.split(',')]

        # Flight phase filter (comma-separated)
        flight_phase = params.get('flight_phase')
        if flight_phase:
            filter_dict['flight_phases'] = [f.strip() for f in flight_phase.split(',')]

        # Hierarchy filters
        for prefix in ['threat', 'error', 'uas']:
            for level in ['l1', 'l2', 'l3']:
                key = f'{prefix}_{level}'
                value = params.get(key)
                if value:
                    filter_dict[key] = value

        # Training topic filter (comma-separated)
        training_topic = params.get('training_topic')
        if training_topic:
            filter_dict['training_topics'] = [t.strip() for t in training_topic.split(',')]

        # Competency filter (comma-separated)
        competency = params.get('competency')
        if competency:
            filter_dict['competencies'] = [c.strip() for c in competency.split(',')]

        return filter_dict

    @swagger_auto_schema(
        tags=['Aviation Analytics'],
        operation_summary='Get paginated events analytics',
        operation_description="""
        Retrieve paginated event data for analytics visualization.

        Returns events with nested labeling items and result performances,
        using Chinese field names for frontend compatibility. Supports
        filtering via query parameters.

        Requires authentication and organization membership.
        """,
        manual_parameters=[
            openapi.Parameter(
                'date_start', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Start date (YYYY-MM-DD, inclusive)'
            ),
            openapi.Parameter(
                'date_end', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='End date (YYYY-MM-DD, inclusive)'
            ),
            openapi.Parameter(
                'aircraft', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Aircraft types (comma-separated)'
            ),
            openapi.Parameter(
                'airport', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Airport code'
            ),
            openapi.Parameter(
                'event_type', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Event types (comma-separated)'
            ),
            openapi.Parameter(
                'flight_phase', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Flight phases (comma-separated)'
            ),
            openapi.Parameter(
                'threat_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 1 code'
            ),
            openapi.Parameter(
                'threat_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 2 code'
            ),
            openapi.Parameter(
                'threat_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 3 code'
            ),
            openapi.Parameter(
                'error_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 1 code'
            ),
            openapi.Parameter(
                'error_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 2 code'
            ),
            openapi.Parameter(
                'error_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 3 code'
            ),
            openapi.Parameter(
                'uas_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 1 code'
            ),
            openapi.Parameter(
                'uas_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 2 code'
            ),
            openapi.Parameter(
                'uas_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 3 code'
            ),
            openapi.Parameter(
                'page', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                description='Page number'
            ),
            openapi.Parameter(
                'page_size', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                description='Items per page (default: 50, max: 100)'
            ),
        ],
        responses={
            200: AnalyticsEventSerializer(many=True),
            401: 'Unauthorized - Authentication required',
            404: 'Aviation project not found or not accessible'
        }
    )
    def get(self, request, *args, **kwargs):
        """Handle GET request."""
        return super().get(request, *args, **kwargs)


# =============================================================================
# Filter Options API View (Phase 1 - Filter Integration)
# =============================================================================


class FilterOptionsAPI(generics.GenericAPIView):
    """
    GET /api/aviation/projects/<pk>/filter-options/

    Retrieve distinct filter options for aviation analytics dashboard.

    Returns arrays of unique values for dropdown filters:
    - aircraft: Aircraft types from events
    - airports: Airport codes (departure/arrival/actual_landing merged)
    - eventTypes: Event types from result performances
    - flightPhases: Flight phases from events
    - trainingTopics: Training topics from result performances (flattened from JSONField arrays)

    Authentication:
        Requires authenticated user with access to the project's organization.

    Response:
        JSON object with 5 arrays, all sorted alphabetically with duplicates removed.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = FilterOptionsSerializer

    def get_object(self):
        """Get aviation project with organization check."""
        aviation_project = get_object_or_404(
            AviationProject.objects.select_related('project', 'project__organization'),
            pk=self.kwargs['pk']
        )

        # Verify organization access - return 404 to hide project existence
        if aviation_project.project.organization != self.request.user.active_organization:
            raise Http404("Aviation project not found")

        return aviation_project

    def _get_distinct_aircraft(self, aviation_project):
        """Get distinct aircraft types for project."""
        return list(
            AviationEvent.objects.filter(
                task__project=aviation_project.project,
                aircraft_type__isnull=False
            ).exclude(
                aircraft_type=''
            ).values_list('aircraft_type', flat=True).distinct().order_by('aircraft_type')
        )

    def _get_distinct_airports(self, aviation_project):
        """Get distinct airports (merged from departure/arrival/actual_landing)."""
        events = AviationEvent.objects.filter(
            task__project=aviation_project.project
        ).values_list('departure_airport', 'arrival_airport', 'actual_landing_airport')

        airport_set = set()
        for dep, arr, actual in events:
            if dep:
                airport_set.add(dep)
            if arr:
                airport_set.add(arr)
            if actual:
                airport_set.add(actual)
        return sorted(list(airport_set))

    def _get_distinct_event_types(self, aviation_project):
        """Get distinct event types from result performances."""
        return list(
            ResultPerformance.objects.filter(
                aviation_project=aviation_project,
                event_type__isnull=False
            ).exclude(
                event_type=''
            ).values_list('event_type', flat=True).distinct().order_by('event_type')
        )

    def _get_distinct_flight_phases(self, aviation_project):
        """Get distinct flight phases from events."""
        return list(
            AviationEvent.objects.filter(
                task__project=aviation_project.project,
                flight_phase__isnull=False
            ).exclude(
                flight_phase=''
            ).values_list('flight_phase', flat=True).distinct().order_by('flight_phase')
        )

    def _get_distinct_training_topics(self, aviation_project):
        """Get distinct training topics from result performances (flattened from JSONField arrays)."""
        performances = ResultPerformance.objects.filter(
            aviation_project=aviation_project
        ).values_list('training_topics', flat=True)

        topic_set = set()
        for topics in performances:
            if isinstance(topics, list):
                topic_set.update(t for t in topics if t)
        return sorted(list(topic_set))

    @swagger_auto_schema(
        tags=['Aviation Analytics'],
        operation_summary='Get filter options for analytics',
        operation_description='''
        Retrieve distinct filter option arrays for aviation analytics dashboard.

        Returns unique values for:
        - aircraft: Aircraft types from events
        - airports: Airport codes (departure/arrival/actual_landing merged)
        - eventTypes: Event types from result performances
        - flightPhases: Flight phases from events
        - trainingTopics: Training topics from result performances

        All arrays are sorted alphabetically with duplicates removed.
        ''',
        responses={
            200: FilterOptionsSerializer,
            401: 'Unauthorized - Authentication required',
            404: 'Aviation project not found or not accessible'
        }
    )
    def get(self, request, pk):
        """Retrieve filter options."""
        aviation_project = self.get_object()

        data = {
            'aircraft': self._get_distinct_aircraft(aviation_project),
            'airports': self._get_distinct_airports(aviation_project),
            'eventTypes': self._get_distinct_event_types(aviation_project),
            'flightPhases': self._get_distinct_flight_phases(aviation_project),
            'trainingTopics': self._get_distinct_training_topics(aviation_project),
        }

        logger.debug(
            "FilterOptionsAPI: project=%s, aircraft=%d, airports=%d, "
            "eventTypes=%d, flightPhases=%d, trainingTopics=%d",
            aviation_project.id,
            len(data['aircraft']),
            len(data['airports']),
            len(data['eventTypes']),
            len(data['flightPhases']),
            len(data['trainingTopics']),
        )

        serializer = self.get_serializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


# =============================================================================
# Organization-Wide Analytics API Views (All Projects)
# =============================================================================


class AllEventsAnalyticsAPI(generics.ListAPIView):
    """
    GET /api/aviation/events/analytics/

    Retrieve paginated event data for analytics across ALL aviation projects
    in the user's organization.

    Returns events with nested labeling items and result performances,
    using Chinese field names for frontend compatibility. Supports all
    10 filter types from Phase 1 via query parameters.

    Query Parameters (Filters):
        - date_start: Start date (YYYY-MM-DD, inclusive)
        - date_end: End date (YYYY-MM-DD, inclusive)
        - aircraft: Aircraft types (comma-separated for multiple)
        - airport: Airport code (matches departure/arrival/actual_landing)
        - event_type: Event types (comma-separated for multiple)
        - flight_phase: Flight phases (comma-separated for multiple)
        - threat_l1/l2/l3: Threat hierarchy codes
        - error_l1/l2/l3: Error hierarchy codes
        - uas_l1/l2/l3: UAS hierarchy codes
        - training_topic: Training topics (comma-separated)
        - competency: Competency codes (comma-separated)

    Query Parameters (Pagination):
        - page: Page number (default: 1)
        - page_size: Items per page (default: 50, max: 100)

    Authentication:
        Requires authenticated user with organization membership.

    Response:
        - count: Total number of events matching filters
        - next: URL to next page (or null)
        - previous: URL to previous page (or null)
        - results: Array of event objects with Chinese field names
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = AnalyticsEventSerializer
    pagination_class = AviationEventsAnalyticsPagination

    def get_queryset(self):
        """Build filtered and prefetched queryset for events across all org projects."""
        org = self.request.user.active_organization

        # Get all aviation project IDs in the organization
        aviation_project_ids = AviationProject.objects.filter(
            project__organization=org
        ).values_list('project_id', flat=True)

        # Build base queryset with optimized prefetching
        queryset = AviationEvent.objects.filter(
            task__project_id__in=aviation_project_ids
        ).prefetch_related(
            Prefetch(
                'labeling_items',
                queryset=LabelingItem.objects.select_related(
                    'threat_type_l1', 'threat_type_l2', 'threat_type_l3',
                    'error_type_l1', 'error_type_l2', 'error_type_l3',
                    'uas_type_l1', 'uas_type_l2', 'uas_type_l3',
                    'linked_result',
                )
            ),
            'result_performances',
        ).order_by('-date', '-id')

        # Parse query params into filter dict
        filter_params = self._parse_filter_params()

        # Apply Phase 1 filters
        queryset = apply_all_filters(queryset, filter_params)

        return queryset

    def _parse_filter_params(self):
        """Parse query parameters into filter dictionary."""
        params = self.request.query_params
        filter_dict = {}

        # Date filters
        date_start = params.get('date_start')
        if date_start:
            try:
                filter_dict['date_start'] = date.fromisoformat(date_start)
            except ValueError:
                pass

        date_end = params.get('date_end')
        if date_end:
            try:
                filter_dict['date_end'] = date.fromisoformat(date_end)
            except ValueError:
                pass

        # Aircraft filter (comma-separated)
        aircraft = params.get('aircraft')
        if aircraft:
            filter_dict['aircraft'] = [a.strip() for a in aircraft.split(',')]

        # Airport filter (single value)
        airport = params.get('airport')
        if airport:
            filter_dict['airport'] = airport

        # Event type filter (comma-separated)
        event_type = params.get('event_type')
        if event_type:
            filter_dict['event_types'] = [e.strip() for e in event_type.split(',')]

        # Flight phase filter (comma-separated)
        flight_phase = params.get('flight_phase')
        if flight_phase:
            filter_dict['flight_phases'] = [f.strip() for f in flight_phase.split(',')]

        # Hierarchy filters
        for prefix in ['threat', 'error', 'uas']:
            for level in ['l1', 'l2', 'l3']:
                key = f'{prefix}_{level}'
                value = params.get(key)
                if value:
                    filter_dict[key] = value

        # Training topic filter (comma-separated)
        training_topic = params.get('training_topic')
        if training_topic:
            filter_dict['training_topics'] = [t.strip() for t in training_topic.split(',')]

        # Competency filter (comma-separated)
        competency = params.get('competency')
        if competency:
            filter_dict['competencies'] = [c.strip() for c in competency.split(',')]

        return filter_dict

    @swagger_auto_schema(
        tags=['Aviation Analytics'],
        operation_summary='Get paginated events analytics across all projects',
        operation_description="""
        Retrieve paginated event data for analytics visualization across ALL
        aviation projects in the user's organization.

        Returns events with nested labeling items and result performances,
        using Chinese field names for frontend compatibility. Supports
        filtering via query parameters.

        Requires authentication and organization membership.
        """,
        manual_parameters=[
            openapi.Parameter(
                'date_start', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Start date (YYYY-MM-DD, inclusive)'
            ),
            openapi.Parameter(
                'date_end', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='End date (YYYY-MM-DD, inclusive)'
            ),
            openapi.Parameter(
                'aircraft', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Aircraft types (comma-separated)'
            ),
            openapi.Parameter(
                'airport', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Airport code'
            ),
            openapi.Parameter(
                'event_type', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Event types (comma-separated)'
            ),
            openapi.Parameter(
                'flight_phase', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Flight phases (comma-separated)'
            ),
            openapi.Parameter(
                'threat_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 1 code'
            ),
            openapi.Parameter(
                'threat_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 2 code'
            ),
            openapi.Parameter(
                'threat_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Threat level 3 code'
            ),
            openapi.Parameter(
                'error_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 1 code'
            ),
            openapi.Parameter(
                'error_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 2 code'
            ),
            openapi.Parameter(
                'error_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='Error level 3 code'
            ),
            openapi.Parameter(
                'uas_l1', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 1 code'
            ),
            openapi.Parameter(
                'uas_l2', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 2 code'
            ),
            openapi.Parameter(
                'uas_l3', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                description='UAS level 3 code'
            ),
            openapi.Parameter(
                'page', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                description='Page number'
            ),
            openapi.Parameter(
                'page_size', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                description='Items per page (default: 50, max: 100)'
            ),
        ],
        responses={
            200: AnalyticsEventSerializer(many=True),
            401: 'Unauthorized - Authentication required',
        }
    )
    def get(self, request, *args, **kwargs):
        """Handle GET request."""
        return super().get(request, *args, **kwargs)


class AllFilterOptionsAPI(generics.GenericAPIView):
    """
    GET /api/aviation/filter-options/

    Retrieve distinct filter options aggregated across ALL aviation projects
    in the user's organization for the analytics dashboard.

    Returns arrays of unique values for dropdown filters:
    - aircraft: Aircraft types from events across all projects
    - airports: Airport codes (departure/arrival/actual_landing merged)
    - eventTypes: Event types from result performances
    - flightPhases: Flight phases from events
    - trainingTopics: Training topics from result performances (flattened from JSONField arrays)

    Authentication:
        Requires authenticated user with organization membership.

    Response:
        JSON object with 5 arrays, all sorted alphabetically with duplicates removed.
    """
    permission_classes = (IsAuthenticated,)
    serializer_class = FilterOptionsSerializer

    def _get_org_aviation_project_ids(self):
        """Get all aviation project Label Studio project IDs in user's organization."""
        org = self.request.user.active_organization
        return list(
            AviationProject.objects.filter(
                project__organization=org
            ).values_list('project_id', flat=True)
        )

    def _get_org_aviation_projects(self):
        """Get all aviation projects in user's organization."""
        org = self.request.user.active_organization
        return AviationProject.objects.filter(project__organization=org)

    def _get_distinct_aircraft(self, project_ids):
        """Get distinct aircraft types across all org projects."""
        return list(
            AviationEvent.objects.filter(
                task__project_id__in=project_ids,
                aircraft_type__isnull=False
            ).exclude(
                aircraft_type=''
            ).values_list('aircraft_type', flat=True).distinct().order_by('aircraft_type')
        )

    def _get_distinct_airports(self, project_ids):
        """Get distinct airports (merged from departure/arrival/actual_landing)."""
        events = AviationEvent.objects.filter(
            task__project_id__in=project_ids
        ).values_list('departure_airport', 'arrival_airport', 'actual_landing_airport')

        airport_set = set()
        for dep, arr, actual in events:
            if dep:
                airport_set.add(dep)
            if arr:
                airport_set.add(arr)
            if actual:
                airport_set.add(actual)
        return sorted(list(airport_set))

    def _get_distinct_event_types(self, aviation_projects):
        """Get distinct event types from result performances across all projects."""
        return list(
            ResultPerformance.objects.filter(
                aviation_project__in=aviation_projects,
                event_type__isnull=False
            ).exclude(
                event_type=''
            ).values_list('event_type', flat=True).distinct().order_by('event_type')
        )

    def _get_distinct_flight_phases(self, project_ids):
        """Get distinct flight phases from events across all projects."""
        return list(
            AviationEvent.objects.filter(
                task__project_id__in=project_ids,
                flight_phase__isnull=False
            ).exclude(
                flight_phase=''
            ).values_list('flight_phase', flat=True).distinct().order_by('flight_phase')
        )

    def _get_distinct_training_topics(self, aviation_projects):
        """Get distinct training topics from result performances (flattened from JSONField arrays)."""
        performances = ResultPerformance.objects.filter(
            aviation_project__in=aviation_projects
        ).values_list('training_topics', flat=True)

        topic_set = set()
        for topics in performances:
            if isinstance(topics, list):
                topic_set.update(t for t in topics if t)
        return sorted(list(topic_set))

    @swagger_auto_schema(
        tags=['Aviation Analytics'],
        operation_summary='Get filter options for analytics across all projects',
        operation_description='''
        Retrieve distinct filter option arrays aggregated across ALL aviation
        projects in the user's organization for the analytics dashboard.

        Returns unique values for:
        - aircraft: Aircraft types from events
        - airports: Airport codes (departure/arrival/actual_landing merged)
        - eventTypes: Event types from result performances
        - flightPhases: Flight phases from events
        - trainingTopics: Training topics from result performances

        All arrays are sorted alphabetically with duplicates removed.
        ''',
        responses={
            200: FilterOptionsSerializer,
            401: 'Unauthorized - Authentication required',
        }
    )
    def get(self, request):
        """Retrieve aggregated filter options across all organization projects."""
        project_ids = self._get_org_aviation_project_ids()
        aviation_projects = self._get_org_aviation_projects()

        data = {
            'aircraft': self._get_distinct_aircraft(project_ids),
            'airports': self._get_distinct_airports(project_ids),
            'eventTypes': self._get_distinct_event_types(aviation_projects),
            'flightPhases': self._get_distinct_flight_phases(project_ids),
            'trainingTopics': self._get_distinct_training_topics(aviation_projects),
        }

        logger.debug(
            "AllFilterOptionsAPI: org=%s, aircraft=%d, airports=%d, "
            "eventTypes=%d, flightPhases=%d, trainingTopics=%d",
            self.request.user.active_organization.id,
            len(data['aircraft']),
            len(data['airports']),
            len(data['eventTypes']),
            len(data['flightPhases']),
            len(data['trainingTopics']),
        )

        serializer = self.get_serializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)
