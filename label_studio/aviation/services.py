import logging
from io import BytesIO

from openpyxl import Workbook

logger = logging.getLogger(__name__)


class AviationExportService:
    """Export aviation project data to JSON or Excel format."""

    def __init__(self, aviation_project_id: int):
        self.aviation_project_id = aviation_project_id

    def _get_aviation_project(self):
        from .models import AviationProject
        return AviationProject.objects.select_related('project').get(id=self.aviation_project_id)

    def _get_events_queryset(self):
        from .models import AviationEvent
        return AviationEvent.objects.filter(
            task__project__aviation_project__id=self.aviation_project_id
        ).select_related(
            'task'
        ).prefetch_related(
            'labeling_items__created_by',
            'labeling_items__reviewed_by',
            'labeling_items__threat_type_l1',
            'labeling_items__threat_type_l2',
            'labeling_items__threat_type_l3',
            'labeling_items__error_type_l1',
            'labeling_items__error_type_l2',
            'labeling_items__error_type_l3',
            'labeling_items__uas_type_l1',
            'labeling_items__uas_type_l2',
            'labeling_items__uas_type_l3',
        ).order_by('date', 'event_number')

    def _get_result_performances_queryset(self):
        from .models import ResultPerformance
        return ResultPerformance.objects.filter(
            aviation_project_id=self.aviation_project_id
        ).select_related(
            'created_by',
            'reviewed_by'
        ).prefetch_related(
            'labeling_item_links__labeling_item'
        ).order_by('created_at')

    def _serialize_type_hierarchy(self, th):
        if th is None:
            return None
        return {'code': th.code, 'label': th.label}

    def _serialize_labeling_item(self, item):
        return {
            'id': item.id,
            'sequence_number': item.sequence_number,
            'status': item.status,
            'threat': {
                'l1': self._serialize_type_hierarchy(item.threat_type_l1),
                'l2': self._serialize_type_hierarchy(item.threat_type_l2),
                'l3': self._serialize_type_hierarchy(item.threat_type_l3),
                'management': item.threat_management,
                'impact': item.threat_impact,
                'coping_abilities': item.threat_coping_abilities,
                'description': item.threat_description,
            },
            'error': {
                'l1': self._serialize_type_hierarchy(item.error_type_l1),
                'l2': self._serialize_type_hierarchy(item.error_type_l2),
                'l3': self._serialize_type_hierarchy(item.error_type_l3),
                'relevance': item.error_relevance,
                'management': item.error_management,
                'impact': item.error_impact,
                'coping_abilities': item.error_coping_abilities,
                'description': item.error_description,
            },
            'uas': {
                'applicable': item.uas_applicable,
                'relevance': item.uas_relevance,
                'l1': self._serialize_type_hierarchy(item.uas_type_l1),
                'l2': self._serialize_type_hierarchy(item.uas_type_l2),
                'l3': self._serialize_type_hierarchy(item.uas_type_l3),
                'management': item.uas_management,
                'impact': item.uas_impact,
                'coping_abilities': item.uas_coping_abilities,
                'description': item.uas_description,
            },
            'calculated_topics': {
                'threat': item.calculated_threat_topics,
                'error': item.calculated_error_topics,
                'uas': item.calculated_uas_topics,
            },
            'notes': item.notes,
            'created_by': item.created_by.username if item.created_by else None,
            'reviewed_by': item.reviewed_by.username if item.reviewed_by else None,
            'created_at': item.created_at.isoformat(),
            'updated_at': item.updated_at.isoformat(),
        }

    def _serialize_event(self, event):
        return {
            'id': event.id,
            'event_number': event.event_number,
            'event_description': event.event_description,
            'date': str(event.date),
            'time': str(event.time) if event.time else None,
            'location': event.location,
            'airport': event.airport,
            'departure_airport': event.departure_airport,
            'arrival_airport': event.arrival_airport,
            'actual_landing_airport': event.actual_landing_airport,
            'flight_phase': event.flight_phase,
            'aircraft_registration': event.aircraft_registration,
            'aircraft_type': event.aircraft_type,
            'crew_composition': event.crew_composition,
            'weather_conditions': event.weather_conditions,
            'labeling_items': [
                self._serialize_labeling_item(item)
                for item in event.labeling_items.all()
            ],
        }

    def _serialize_result_performance(self, rp):
        linked_items = []
        for link in rp.labeling_item_links.all():
            linked_items.append({
                'labeling_item_id': link.labeling_item_id,
                'contribution_weight': str(link.contribution_weight),
                'notes': link.notes,
            })

        return {
            'id': rp.id,
            'event_type': rp.event_type,
            'flight_phase': rp.flight_phase,
            'likelihood': rp.likelihood,
            'severity': rp.severity,
            'training_effect': rp.training_effect,
            'training_plan': rp.training_plan,
            'training_topics': rp.training_topics,
            'training_goals': rp.training_goals,
            'objectives': rp.objectives,
            'recommendations': rp.recommendations,
            'status': rp.status,
            'created_by': rp.created_by.username if rp.created_by else None,
            'reviewed_by': rp.reviewed_by.username if rp.reviewed_by else None,
            'linked_items': linked_items,
            'created_at': rp.created_at.isoformat(),
            'updated_at': rp.updated_at.isoformat(),
        }

    def export_to_json(self) -> dict:
        from django.utils import timezone

        aviation_project = self._get_aviation_project()
        events = list(self._get_events_queryset())
        result_performances = list(self._get_result_performances_queryset())

        return {
            'metadata': {
                'export_date': timezone.now().isoformat(),
                'project_id': aviation_project.project_id,
                'project_title': aviation_project.project.title,
                'total_events': len(events),
            },
            'events': [self._serialize_event(event) for event in events],
            'result_performances': [
                self._serialize_result_performance(rp) for rp in result_performances
            ],
        }

    def export_to_xlsx(self) -> BytesIO:
        events = list(self._get_events_queryset())
        result_performances = list(self._get_result_performances_queryset())

        wb = Workbook()

        ws_events = wb.active
        ws_events.title = 'Events'
        event_headers = [
            'id', 'event_number', 'event_description', 'date', 'time',
            'location', 'airport', 'departure_airport', 'arrival_airport',
            'actual_landing_airport', 'flight_phase', 'aircraft_registration',
            'aircraft_type', 'weather_conditions'
        ]
        ws_events.append(event_headers)
        for event in events:
            ws_events.append([
                event.id,
                event.event_number,
                event.event_description,
                str(event.date),
                str(event.time) if event.time else None,
                event.location,
                event.airport,
                event.departure_airport,
                event.arrival_airport,
                event.actual_landing_airport,
                event.flight_phase,
                event.aircraft_registration,
                event.aircraft_type,
                event.weather_conditions,
            ])

        ws_items = wb.create_sheet('LabelingItems')
        item_headers = [
            'id', 'event_number', 'sequence_number', 'status',
            'threat_l1_code', 'threat_l2_code', 'threat_l3_code', 'threat_description',
            'error_l1_code', 'error_l2_code', 'error_l3_code', 'error_description',
            'uas_applicable', 'uas_l1_code', 'uas_l2_code', 'uas_l3_code', 'uas_description',
            'notes', 'created_by', 'reviewed_by', 'created_at', 'updated_at'
        ]
        ws_items.append(item_headers)
        for event in events:
            for item in event.labeling_items.all():
                ws_items.append([
                    item.id,
                    event.event_number,
                    item.sequence_number,
                    item.status,
                    item.threat_type_l1.code if item.threat_type_l1 else None,
                    item.threat_type_l2.code if item.threat_type_l2 else None,
                    item.threat_type_l3.code if item.threat_type_l3 else None,
                    item.threat_description,
                    item.error_type_l1.code if item.error_type_l1 else None,
                    item.error_type_l2.code if item.error_type_l2 else None,
                    item.error_type_l3.code if item.error_type_l3 else None,
                    item.error_description,
                    item.uas_applicable,
                    item.uas_type_l1.code if item.uas_type_l1 else None,
                    item.uas_type_l2.code if item.uas_type_l2 else None,
                    item.uas_type_l3.code if item.uas_type_l3 else None,
                    item.uas_description,
                    item.notes,
                    item.created_by.username if item.created_by else None,
                    item.reviewed_by.username if item.reviewed_by else None,
                    item.created_at.isoformat(),
                    item.updated_at.isoformat(),
                ])

        ws_rp = wb.create_sheet('ResultPerformances')
        rp_headers = [
            'id', 'event_type', 'flight_phase', 'likelihood', 'severity',
            'training_effect', 'training_plan', 'training_goals',
            'objectives', 'recommendations', 'status',
            'created_by', 'reviewed_by', 'created_at', 'updated_at'
        ]
        ws_rp.append(rp_headers)
        for rp in result_performances:
            ws_rp.append([
                rp.id,
                rp.event_type,
                rp.flight_phase,
                rp.likelihood,
                rp.severity,
                rp.training_effect,
                rp.training_plan,
                rp.training_goals,
                rp.objectives,
                rp.recommendations,
                rp.status,
                rp.created_by.username if rp.created_by else None,
                rp.reviewed_by.username if rp.reviewed_by else None,
                rp.created_at.isoformat(),
                rp.updated_at.isoformat(),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# =============================================================================
# Review Notification Service
# =============================================================================

def send_review_notification(review_decision):
    """
    Send notification to annotator when their work is reviewed.

    Args:
        review_decision: ReviewDecision instance with status and related labeling_item

    Notification is sent to labeling_item.created_by (the annotator).
    Skips silently if no recipient (created_by is null).
    """
    from django.utils import timezone
    from notifications.models import NotificationChannel, NotificationEventType
    from notifications.services import NotificationService

    labeling_item = review_decision.labeling_item
    recipient = labeling_item.created_by

    # Skip if no recipient
    if not recipient:
        logger.debug(f'No recipient for review notification on item {labeling_item.id}')
        return

    # Determine event type and message based on status
    status_map = {
        'approved': {
            'event_type': NotificationEventType.REVIEW_APPROVED,
            'action': 'approved',
        },
        'rejected_partial': {
            'event_type': NotificationEventType.REVIEW_REJECTED,
            'action': 'rejected (partial)',
        },
        'rejected_full': {
            'event_type': NotificationEventType.REVIEW_REJECTED,
            'action': 'rejected',
        },
        'revision_requested': {
            'event_type': NotificationEventType.REVIEW_REVISION_REQUESTED,
            'action': 'requested revision on',
        },
    }

    status_info = status_map.get(review_decision.status)
    if not status_info:
        logger.warning(f'Unknown review status: {review_decision.status}')
        return

    # Get event number for subject
    event = labeling_item.event
    event_number = event.event_number

    # Get aviation project ID for path
    try:
        aviation_project = event.task.project.aviation_project
        aviation_project_id = aviation_project.id
    except AttributeError:
        logger.warning(f'Could not find aviation project for event {event.id}')
        aviation_project_id = 0

    # Build reviewer name
    reviewer = review_decision.reviewer
    reviewer_name = reviewer.email if reviewer else 'A reviewer'

    # Build notification data
    subject = f'Review Decision: {event_number}'
    message = f'{reviewer_name} has {status_info["action"]} your annotation'
    path = f'/aviation/projects/{aviation_project_id}/events/{event.id}?item={labeling_item.id}'

    # Send the notification synchronously
    notification_service = NotificationService()

    try:
        notification_service.send_notification_sync(
            channel_name=NotificationChannel.NOTIFICATION,
            event_type=status_info['event_type'],
            subject=subject,
            message=message,
            ts=timezone.now(),
            receive_user=recipient,
            path=path,
            action_type=review_decision.status,
            source='aviation_review'
        )
    except Exception as e:
        logger.error(f'Failed to send review notification: {e}')


def send_resubmit_notification(labeling_item, resubmitter):
    """
    Send notification to reviewer when item is resubmitted.

    Args:
        labeling_item: LabelingItem instance being resubmitted
        resubmitter: User who resubmitted the item (usually the annotator)

    Notification is sent to the last reviewer (from most recent ReviewDecision).
    Skips silently if no reviewer found.
    """
    from django.utils import timezone
    from notifications.models import NotificationChannel, NotificationEventType
    from notifications.services import NotificationService
    from .models import ReviewDecision

    # Find the last reviewer
    last_decision = ReviewDecision.objects.filter(
        labeling_item=labeling_item
    ).order_by('-created_at').first()

    if not last_decision or not last_decision.reviewer:
        logger.debug(f'No reviewer to notify for resubmit on item {labeling_item.id}')
        return

    recipient = last_decision.reviewer

    # Get event number for subject
    event = labeling_item.event
    event_number = event.event_number

    # Get aviation project ID for path
    try:
        aviation_project = event.task.project.aviation_project
        aviation_project_id = aviation_project.id
    except AttributeError:
        logger.warning(f'Could not find aviation project for event {event.id}')
        aviation_project_id = 0

    # Build resubmitter name
    resubmitter_name = resubmitter.email if resubmitter else 'An annotator'

    # Build notification data
    subject = f'Annotation Resubmitted: {event_number}'
    message = f'{resubmitter_name} has resubmitted annotation for review'
    path = f'/aviation/projects/{aviation_project_id}/events/{event.id}?item={labeling_item.id}'

    # Send the notification synchronously
    notification_service = NotificationService()

    try:
        notification_service.send_notification_sync(
            channel_name=NotificationChannel.NOTIFICATION,
            event_type=NotificationEventType.REVIEW_RESUBMITTED,
            subject=subject,
            message=message,
            ts=timezone.now(),
            receive_user=recipient,
            path=path,
            action_type='resubmitted',
            source='aviation_review'
        )
    except Exception as e:
        logger.error(f'Failed to send resubmit notification: {e}')
