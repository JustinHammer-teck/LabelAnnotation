from io import BytesIO
from django.test import TestCase
from openpyxl import load_workbook

from organizations.tests.factories import OrganizationFactory
from projects.tests.factories import ProjectFactory
from tasks.tests.factories import TaskFactory

from aviation.services import AviationExportService
from aviation.tests.factories import (
    AviationProjectFactory,
    AviationEventFactory,
    TypeHierarchyFactory,
    LabelingItemFactory,
    ResultPerformanceFactory,
    LabelingItemPerformanceFactory,
)


class AviationExportServiceTest(TestCase):
    def setUp(self):
        self.organization = OrganizationFactory()
        self.project = ProjectFactory(organization=self.organization)
        self.aviation_project = AviationProjectFactory(project=self.project)

        self.threat_l1 = TypeHierarchyFactory(
            category='threat',
            level=1,
            code='T1',
            label='Threat Level 1',
        )
        self.threat_l2 = TypeHierarchyFactory(
            category='threat',
            level=2,
            code='T1.1',
            label='Threat Level 2',
            parent=self.threat_l1,
        )
        self.threat_l3 = TypeHierarchyFactory(
            category='threat',
            level=3,
            code='T1.1.1',
            label='Threat Level 3',
            parent=self.threat_l2,
        )

        self.task = TaskFactory(project=self.project)
        self.event = AviationEventFactory(task=self.task, event_number='EVT-00001')

        self.labeling_item = LabelingItemFactory(
            event=self.event,
            threat_type_l1=self.threat_l1,
            threat_type_l2=self.threat_l2,
            threat_type_l3=self.threat_l3,
        )

        self.result_performance = ResultPerformanceFactory(
            aviation_project=self.aviation_project
        )
        self.labeling_item_performance = LabelingItemPerformanceFactory(
            labeling_item=self.labeling_item,
            result_performance=self.result_performance,
        )

    def test_export_to_json_returns_valid_structure(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_json()

        self.assertIn('metadata', result)
        self.assertIn('events', result)
        self.assertIn('result_performances', result)

    def test_export_to_json_includes_all_events(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_json()

        self.assertEqual(len(result['events']), 1)
        self.assertEqual(result['events'][0]['event_number'], 'EVT-00001')

    def test_export_to_json_includes_labeling_items(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_json()

        self.assertIn('labeling_items', result['events'][0])
        self.assertEqual(len(result['events'][0]['labeling_items']), 1)

    def test_export_to_json_resolves_type_hierarchy_codes(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_json()

        item = result['events'][0]['labeling_items'][0]
        self.assertEqual(item['threat']['l1']['code'], 'T1')
        self.assertEqual(item['threat']['l1']['label'], 'Threat Level 1')

    def test_export_to_xlsx_returns_valid_workbook(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_xlsx()

        self.assertIsInstance(result, BytesIO)
        workbook = load_workbook(result)
        self.assertIsNotNone(workbook)

    def test_export_to_xlsx_has_correct_sheets(self):
        service = AviationExportService(self.aviation_project.id)
        result = service.export_to_xlsx()

        workbook = load_workbook(result)
        sheet_names = workbook.sheetnames

        self.assertIn('Events', sheet_names)
        self.assertIn('LabelingItems', sheet_names)
        self.assertIn('ResultPerformances', sheet_names)

    def test_export_empty_project(self):
        empty_project = ProjectFactory(organization=self.organization)
        empty_aviation_project = AviationProjectFactory(project=empty_project)

        service = AviationExportService(empty_aviation_project.id)
        result = service.export_to_json()

        self.assertEqual(result['events'], [])
        self.assertEqual(result['result_performances'], [])

    def test_queryset_optimization(self):
        for i in range(5):
            task = TaskFactory(project=self.project)
            event = AviationEventFactory(task=task, event_number=f'EVT-{i:05d}')
            for j in range(3):
                LabelingItemFactory(
                    event=event,
                    sequence_number=j + 1,
                    threat_type_l1=self.threat_l1,
                    threat_type_l2=self.threat_l2,
                    threat_type_l3=self.threat_l3,
                )

        service = AviationExportService(self.aviation_project.id)
        with self.assertNumQueries(9):
            service.export_to_json()
